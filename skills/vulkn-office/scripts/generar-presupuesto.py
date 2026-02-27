#!/usr/bin/env python3
"""
Generador de Presupuestos en Excel

Uso:
    python generar-presupuesto.py --titulo "Presupuesto Q1 2026" \
        --items '[{"concepto":"Nómina","monto":50000},{"concepto":"Renta","monto":15000}]' \
        --output presupuesto.xlsx

    # Con categorías
    python generar-presupuesto.py --titulo "Presupuesto Anual" \
        --categorias '[
            {"nombre":"Gastos Fijos","items":[{"concepto":"Renta","monto":15000}]},
            {"nombre":"Gastos Variables","items":[{"concepto":"Publicidad","monto":5000}]}
        ]' \
        --output presupuesto.xlsx
"""

import sys
import json
import argparse
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: Instala openpyxl con 'pip install openpyxl'")
    sys.exit(1)


def create_budget(titulo, items=None, categorias=None, output='presupuesto.xlsx'):
    """Crea un presupuesto en Excel"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    
    # Estilos
    header_font = Font(bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    category_fill = PatternFill(start_color='4a4a6a', end_color='4a4a6a', fill_type='solid')
    total_font = Font(bold=True, size=12)
    currency_format = '"$"#,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=18)
    ws.merge_cells('A1:D1')
    
    # Fecha
    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(italic=True, size=10)
    
    # Headers
    row = 4
    headers = ['Concepto', 'Monto Mensual', 'Monto Anual', 'Notas']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row += 1
    start_data_row = row
    
    # Datos
    if categorias:
        # Con categorías
        for cat in categorias:
            # Categoría header
            ws.cell(row=row, column=1, value=cat['nombre']).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = category_fill
            ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            cat_start = row
            for item in cat.get('items', []):
                ws.cell(row=row, column=1, value=item['concepto']).border = thin_border
                ws.cell(row=row, column=2, value=item['monto']).border = thin_border
                ws.cell(row=row, column=2).number_format = currency_format
                ws.cell(row=row, column=3, value=f'=B{row}*12').border = thin_border
                ws.cell(row=row, column=3).number_format = currency_format
                ws.cell(row=row, column=4, value=item.get('notas', '')).border = thin_border
                row += 1
            
            # Subtotal de categoría
            ws.cell(row=row, column=1, value=f'Subtotal {cat["nombre"]}').font = Font(italic=True)
            ws.cell(row=row, column=2, value=f'=SUM(B{cat_start}:B{row-1})')
            ws.cell(row=row, column=2).number_format = currency_format
            ws.cell(row=row, column=2).font = Font(italic=True)
            ws.cell(row=row, column=3, value=f'=SUM(C{cat_start}:C{row-1})')
            ws.cell(row=row, column=3).number_format = currency_format
            ws.cell(row=row, column=3).font = Font(italic=True)
            row += 1
            row += 1  # Espacio entre categorías
    
    elif items:
        # Sin categorías, lista simple
        for item in items:
            ws.cell(row=row, column=1, value=item['concepto']).border = thin_border
            ws.cell(row=row, column=2, value=item['monto']).border = thin_border
            ws.cell(row=row, column=2).number_format = currency_format
            ws.cell(row=row, column=3, value=f'=B{row}*12').border = thin_border
            ws.cell(row=row, column=3).number_format = currency_format
            ws.cell(row=row, column=4, value=item.get('notas', '')).border = thin_border
            row += 1
    
    # Total final
    row += 1
    ws.cell(row=row, column=1, value='TOTAL').font = total_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    
    # Suma de todos los montos en columna B
    ws.cell(row=row, column=2, value=f'=SUMIF(A{start_data_row}:A{row-1},"<>Subtotal*",B{start_data_row}:B{row-1})')
    ws.cell(row=row, column=2).number_format = currency_format
    ws.cell(row=row, column=2).font = total_font
    ws.cell(row=row, column=2).fill = header_fill
    ws.cell(row=row, column=2).font = Font(bold=True, color='FFFFFF')
    
    ws.cell(row=row, column=3, value=f'=B{row}*12')
    ws.cell(row=row, column=3).number_format = currency_format
    ws.cell(row=row, column=3).font = total_font
    ws.cell(row=row, column=3).fill = header_fill
    ws.cell(row=row, column=3).font = Font(bold=True, color='FFFFFF')
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    
    # Guardar
    wb.save(output)
    return output


def main():
    parser = argparse.ArgumentParser(description='Generador de Presupuestos Excel')
    parser.add_argument('--titulo', default='Presupuesto', help='Título del presupuesto')
    parser.add_argument('--items', help='Items JSON: [{"concepto":"X","monto":1000}]')
    parser.add_argument('--categorias', help='Categorías JSON con items anidados')
    parser.add_argument('--output', '-o', default='presupuesto.xlsx', help='Archivo de salida')
    
    args = parser.parse_args()
    
    items = None
    categorias = None
    
    if args.categorias:
        try:
            categorias = json.loads(args.categorias)
        except json.JSONDecodeError as e:
            print(f"Error en JSON de categorías: {e}")
            sys.exit(1)
    elif args.items:
        try:
            items = json.loads(args.items)
        except json.JSONDecodeError as e:
            print(f"Error en JSON de items: {e}")
            sys.exit(1)
    else:
        # Datos de ejemplo
        items = [
            {'concepto': 'Nómina', 'monto': 50000},
            {'concepto': 'Renta', 'monto': 15000},
            {'concepto': 'Servicios', 'monto': 5000},
            {'concepto': 'Insumos', 'monto': 10000},
        ]
    
    output = create_budget(args.titulo, items=items, categorias=categorias, output=args.output)
    print(f"✅ Presupuesto generado: {output}")


if __name__ == '__main__':
    main()
